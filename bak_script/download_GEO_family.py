# -*- encoding:utf-8 -*-
# @time: 2023/5/27 19:46
# @author: ifeng

import os
from concurrent.futures import ThreadPoolExecutor

import requests
from lxml import html
from openpyxl import load_workbook


def download_file(nid, path, filename, url):
    for item in range(3):
        try:
            filepath = os.path.join(os.getcwd(), "download", str(nid), path)
            full_path = os.path.join(filepath, filename)
            if not os.path.exists(filepath):  # 判断文件路径是否存在
                os.makedirs(filepath)
            if not os.path.exists(full_path):  # 判断文件是否存在
                resp = requests.get(url)
                with open(full_path, 'wb') as f:
                    f.write(resp.content)
            print(f'No.{nid} 的 {path}{filename} 抓取完毕')
            return
        except Exception as e:
            print(nid, "出错了, 正在重试 -> ", e)


def get_filename(nid, path, url):
    for item in range(3):
        try:
            detail_url = url + path
            resp = requests.get(detail_url)
            etree = html.etree
            et = etree.HTML(resp.text)
            a_list = et.xpath('//html/body/pre/a')
            for a in a_list[1:]:
                filename = a.xpath('./text()')[0]
                full_url = detail_url + filename
                # 下载文件到指定位置
                download_file(nid, path, filename, full_url)
            return
        except Exception as e:
            print(nid, "出错了, 正在重试 -> ", e)


def get_GEO_family(nid, accession_number):
    for item in range(3):
        try:
            for accession in accession_number.split(','):  # 如果存在多个GSE
                path = accession.replace(accession[-3:], 'nnn')  # 将后面三位替换成nnn
                url = f'https://ftp.ncbi.nlm.nih.gov/geo/series/{path}/{accession}/'
                resp = requests.get(url)
                etree = html.etree
                et = etree.HTML(resp.text)
                a_list = et.xpath('//html/body/pre/a')
                with ThreadPoolExecutor(2) as t:
                    for a in a_list[1:]:
                        path = a.xpath('./text()')[0]
                        # 获取每一个文件夹下的所有文件
                        t.submit(get_filename, nid, path, url)
            return
        except Exception as e:
            print(nid, '出错了, 正在重试 -> ', e)


def get_xlsx_info():
    filename = os.path.join(os.path.dirname(__file__), 'download_all.xlsx')

    if os.path.exists(filename):
        a = load_workbook(filename)  # 后缀为xlsx
        sheet = a.active
        for row in range(2, sheet.max_row + 1):  # 从2开始, 因为第一列不需要
            nid = sheet['A' + str(row)].value  # 第一列
            accession_number = sheet['D' + str(row)].value  # 第四列
            yield nid, accession_number
    else:
        return "download_all.xlsx is no exist,pls check file"


def main():
    gen = get_xlsx_info()
    with ThreadPoolExecutor(10000) as t:
        for nid, accession_number in gen:
            # 获取family的所有文件夹名
            t.submit(get_GEO_family, nid, accession_number)

# if __name__ == '__main__':
#     main()

